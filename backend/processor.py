"""
Procesador de datos de IBOPE para cálculo de inversión en radio.
"""
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from data.config import CPM_CONFIG, TIPO_MAPPING, MES_TO_RANKING, get_ranking


def normalizar_mes(mes: str) -> str:
    """Normaliza el nombre del mes a minúsculas sin tildes."""
    mes_lower = str(mes).lower().strip()
    # Remover tildes comunes
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u'}
    for old, new in replacements.items():
        mes_lower = mes_lower.replace(old, new)
    return mes_lower


def process_ibope_file(file_content: bytes, filename: str) -> BytesIO:
    """
    Procesa un archivo Excel de IBOPE y devuelve un Excel con los cálculos de inversión.
    
    El archivo de entrada debe tener las columnas:
    - MARCA
    - TIPO
    - EMISORA/SITE
    - MES
    - AÑO
    - Suma de SPOTS (o similar)
    """
    # Leer el archivo
    df = pd.read_excel(BytesIO(file_content))
    
    # Normalizar nombres de columnas
    df.columns = [str(col).strip().upper() for col in df.columns]
    
    # Mapear columnas conocidas
    column_mapping = {
        'EMISORA/SITE': 'EMISORA',
        'SUMA DE SPOTS': 'SPOTS',
        'SUMA DE SPOT': 'SPOTS',
        'SPOTS': 'SPOTS',
        'AÑO': 'AÑO',
        'ANO': 'AÑO',
    }
    
    for old_name, new_name in column_mapping.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})
    
    # Verificar columnas requeridas
    required_cols = ['MARCA', 'TIPO', 'EMISORA', 'MES', 'AÑO', 'SPOTS']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Columnas faltantes en el archivo: {missing_cols}. Columnas encontradas: {list(df.columns)}")
    
    # Limpiar datos
    df = df.dropna(subset=['MARCA', 'TIPO', 'EMISORA', 'SPOTS'])
    df['SPOTS'] = pd.to_numeric(df['SPOTS'], errors='coerce').fillna(0).astype(int)
    df['AÑO'] = pd.to_numeric(df['AÑO'], errors='coerce').fillna(2024).astype(int)
    
    # Calcular columnas adicionales
    df['MES_NORM'] = df['MES'].apply(normalizar_mes)
    df['MES_RANKING'] = df['MES_NORM'].map(MES_TO_RANKING).fillna('Marzo')
    df['TIPO_AGRUP'] = df['TIPO'].map(TIPO_MAPPING).fillna('SPOT')
    df['CPM'] = df['TIPO_AGRUP'].map(CPM_CONFIG).fillna(11.6)
    
    # Calcular ranking
    df['RANKING_MILES'] = df.apply(
        lambda row: get_ranking(row['AÑO'], row['MES_RANKING'], row['EMISORA']), 
        axis=1
    )
    
    # Calcular impactos e inversión
    df['IMPACTOS'] = (df['SPOTS'] * df['RANKING_MILES'] * 1000).round(0).astype(int)
    df['INVERSION_SOLES'] = (df['IMPACTOS'] * df['CPM'] / 1000).round(2)
    
    # Crear Excel de salida
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Detalle completo
        df_output = df[['MARCA', 'TIPO', 'TIPO_AGRUP', 'EMISORA', 'MES', 'AÑO', 
                        'MES_RANKING', 'SPOTS', 'CPM', 'RANKING_MILES', 'IMPACTOS', 'INVERSION_SOLES']]
        df_output.to_excel(writer, sheet_name='Detalle', index=False)
        
        # Hoja 2: Resumen por Marca
        resumen_marca = df.groupby('MARCA').agg({
            'SPOTS': 'sum',
            'IMPACTOS': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_marca.columns = ['Marca', 'Total Spots', 'Total Impactos', 'Inversión (S/)']
        resumen_marca.to_excel(writer, sheet_name='Resumen por Marca', index=False)
        
        # Hoja 3: Resumen por Marca y Tipo
        resumen_tipo = df.groupby(['MARCA', 'TIPO_AGRUP']).agg({
            'SPOTS': 'sum',
            'IMPACTOS': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_tipo.columns = ['Marca', 'Tipo', 'Total Spots', 'Total Impactos', 'Inversión (S/)']
        resumen_tipo.to_excel(writer, sheet_name='Resumen por Tipo', index=False)
        
        # Hoja 4: Resumen por Marca y Mes
        resumen_mes = df.groupby(['MARCA', 'AÑO', 'MES']).agg({
            'SPOTS': 'sum',
            'IMPACTOS': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_mes.columns = ['Marca', 'Año', 'Mes', 'Total Spots', 'Total Impactos', 'Inversión (S/)']
        resumen_mes.to_excel(writer, sheet_name='Resumen por Mes', index=False)
        
        # Hoja 5: Resumen por Emisora
        resumen_emisora = df.groupby(['MARCA', 'EMISORA']).agg({
            'SPOTS': 'sum',
            'IMPACTOS': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_emisora.columns = ['Marca', 'Emisora', 'Total Spots', 'Total Impactos', 'Inversión (S/)']
        resumen_emisora.to_excel(writer, sheet_name='Resumen por Emisora', index=False)
        
        # Aplicar formato a todas las hojas
        workbook = writer.book
        
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(bold=True, color='FFFFFF')
        
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Formato de headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output


def get_summary_stats(file_content: bytes) -> dict:
    """
    Devuelve estadísticas resumidas del archivo procesado.
    """
    df = pd.read_excel(BytesIO(file_content))
    
    # Normalizar columnas
    df.columns = [str(col).strip().upper() for col in df.columns]
    column_mapping = {
        'EMISORA/SITE': 'EMISORA',
        'SUMA DE SPOTS': 'SPOTS',
    }
    for old_name, new_name in column_mapping.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})
    
    return {
        'total_filas': len(df),
        'marcas': df['MARCA'].nunique() if 'MARCA' in df.columns else 0,
        'emisoras': df['EMISORA'].nunique() if 'EMISORA' in df.columns else 0,
        'columnas': list(df.columns)
    }
