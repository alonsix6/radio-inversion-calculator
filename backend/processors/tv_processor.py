"""
Procesador de datos de Instar para cálculo de inversión en TV.
"""
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config.tv_config import (
    TIPO_MAPPING_TV,
    get_tv_config
)


def process_tv_file(file_content: bytes, filename: str) -> BytesIO:
    """
    Procesa un archivo Excel de Instar y devuelve un Excel con los cálculos de inversión.

    El archivo de entrada debe tener las columnas:
    - MARCA
    - TIPO COMERCIAL
    - CANAL
    - MES
    - AÑO
    - TOTAL SPOTS
    - DURACIÓN (segundos)
    - GRP# [RATING] (rating agregado por fila)
    """
    df = pd.read_excel(BytesIO(file_content))

    # Normalizar nombres de columnas
    df.columns = [str(col).strip().upper() for col in df.columns]

    # Mapear columnas conocidas (nombres de Instar)
    column_mapping = {
        # Canal
        'CANAL/SITE': 'CANAL',
        # Spots/Avisos
        'TOTAL SPOTS': 'SPOTS',
        'SUMA DE SPOTS': 'SPOTS',
        'SUMA DE SPOT': 'SPOTS',
        'SUMA DE AVISOS': 'SPOTS',
        'AVISOS': 'SPOTS',
        'SPOTS': 'SPOTS',
        # Tipo
        'TIPO COMERCIAL': 'TIPO',
        # Año
        'AÑO': 'AÑO',
        'ANO': 'AÑO',
        # Duración/Segundos
        'DURACION': 'SEGUNDOS',
        'DURACIÓN': 'SEGUNDOS',
        'SUMA DE SEGUNDOS': 'SEGUNDOS',
        # GRP# (impactos en miles) - para BANNER, P/D
        'GRP# [RATING]': 'GRP_NUM',
        'GRP#': 'GRP_NUM',
        # GRP% (rating porcentaje) - para SPOTS
        'GRP% [RATING]': 'GRP_PCT',
        'GRP%': 'GRP_PCT',
    }

    for old_name, new_name in column_mapping.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})

    # Verificar columnas requeridas
    required_cols = ['MARCA', 'TIPO', 'CANAL', 'MES', 'AÑO', 'SPOTS', 'GRP_NUM', 'GRP_PCT']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Columnas faltantes en el archivo: {missing_cols}. Columnas encontradas: {list(df.columns)}")

    # Verificar columna SEGUNDOS (opcional pero recomendada para SPOTS)
    if 'SEGUNDOS' not in df.columns:
        df['SEGUNDOS'] = 30  # Default de 30 segundos

    # Limpiar datos
    df = df.dropna(subset=['MARCA', 'TIPO', 'CANAL', 'SPOTS'])
    df['SPOTS'] = pd.to_numeric(df['SPOTS'], errors='coerce').fillna(0).astype(int)
    df['AÑO'] = pd.to_numeric(df['AÑO'], errors='coerce').fillna(2024).astype(int)
    df['SEGUNDOS'] = pd.to_numeric(df['SEGUNDOS'], errors='coerce').fillna(30).astype(int)
    df['GRP_NUM'] = pd.to_numeric(df['GRP_NUM'], errors='coerce').fillna(0)  # Impactos en miles
    df['GRP_PCT'] = pd.to_numeric(df['GRP_PCT'], errors='coerce').fillna(0)  # Rating %

    # Calcular columnas adicionales
    df['TIPO_AGRUP'] = df['TIPO'].str.upper().map(TIPO_MAPPING_TV).fillna('SPOT')

    # Obtener CPM/CPR para cada fila
    def get_cpm_cpr(row):
        config = get_tv_config(row['CANAL'], row['TIPO_AGRUP'])
        return pd.Series({'CPM': config['cpm'], 'CPR': config['cpr']})

    df[['CPM', 'CPR']] = df.apply(get_cpm_cpr, axis=1)

    # Calcular impactos usando GRP# (impactos en miles) - para referencia
    # Impactos = GRP# × 1000 (convertir a personas)
    df['Impactos_Miles'] = (df['GRP_NUM'] * 1000).round(0).astype(int)

    # Calcular inversión según tipo
    # Para SPOTS: GRP% × CPR × Segundos (rating porcentaje)
    # Para otros: GRP# × CPM (impactos en miles)
    def calcular_inversion(row):
        if row['TIPO_AGRUP'] == 'SPOT' or row['TIPO_AGRUP'] == 'NOTICIEROS':
            # Usar GRP% (rating porcentaje) × CPR × Segundos
            return round(row['GRP_PCT'] * row['CPR'] * row['SEGUNDOS'], 2)
        else:
            # Usar GRP# (impactos en miles) × CPM
            return round(row['GRP_NUM'] * row['CPM'], 2)

    df['INVERSION_SOLES'] = df.apply(calcular_inversion, axis=1)

    # Crear Excel de salida
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Hoja 1: Detalle completo
        df_output = df[['MARCA', 'TIPO', 'TIPO_AGRUP', 'CANAL', 'MES', 'AÑO',
                        'SPOTS', 'SEGUNDOS', 'GRP_NUM', 'GRP_PCT', 'CPM', 'CPR',
                        'Impactos_Miles', 'INVERSION_SOLES']]
        df_output.to_excel(writer, sheet_name='Detalle', index=False)

        # Hoja 2: Resumen por Marca
        resumen_marca = df.groupby('MARCA').agg({
            'SPOTS': 'sum',
            'GRP_NUM': 'sum',
            'GRP_PCT': 'sum',
            'Impactos_Miles': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_marca.columns = ['Marca', 'Total Avisos', 'GRP#', 'GRP%', 'Impactos_Miles', 'Inversión (S/)']
        resumen_marca.to_excel(writer, sheet_name='Resumen por Marca', index=False)

        # Hoja 3: Resumen por Marca y Tipo
        resumen_tipo = df.groupby(['MARCA', 'TIPO_AGRUP']).agg({
            'SPOTS': 'sum',
            'GRP_NUM': 'sum',
            'GRP_PCT': 'sum',
            'Impactos_Miles': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_tipo.columns = ['Marca', 'Tipo', 'Total Avisos', 'GRP#', 'GRP%', 'Impactos_Miles', 'Inversión (S/)']
        resumen_tipo.to_excel(writer, sheet_name='Resumen por Tipo', index=False)

        # Hoja 4: Resumen por Marca y Mes
        resumen_mes = df.groupby(['MARCA', 'AÑO', 'MES']).agg({
            'SPOTS': 'sum',
            'GRP_NUM': 'sum',
            'GRP_PCT': 'sum',
            'Impactos_Miles': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_mes.columns = ['Marca', 'Año', 'Mes', 'Total Avisos', 'GRP#', 'GRP%', 'Impactos_Miles', 'Inversión (S/)']
        resumen_mes.to_excel(writer, sheet_name='Resumen por Mes', index=False)

        # Hoja 5: Resumen por Canal
        resumen_canal = df.groupby(['MARCA', 'CANAL']).agg({
            'SPOTS': 'sum',
            'GRP_NUM': 'sum',
            'GRP_PCT': 'sum',
            'Impactos_Miles': 'sum',
            'INVERSION_SOLES': 'sum'
        }).reset_index()
        resumen_canal.columns = ['Marca', 'Canal', 'Total Avisos', 'GRP#', 'GRP%', 'Impactos_Miles', 'Inversión (S/)']
        resumen_canal.to_excel(writer, sheet_name='Resumen por Canal', index=False)

        # Aplicar formato a todas las hojas
        _apply_excel_formatting(writer.book)

    output.seek(0)
    return output


def get_tv_summary_stats(file_content: bytes) -> dict:
    """
    Devuelve estadísticas resumidas del archivo de TV procesado.
    """
    df = pd.read_excel(BytesIO(file_content))

    df.columns = [str(col).strip().upper() for col in df.columns]
    column_mapping = {
        'CANAL/SITE': 'CANAL',
        'TOTAL SPOTS': 'SPOTS',
        'SUMA DE SPOTS': 'SPOTS',
        'SUMA DE AVISOS': 'SPOTS',
    }
    for old_name, new_name in column_mapping.items():
        if old_name in df.columns:
            df = df.rename(columns={old_name: new_name})

    return {
        'total_filas': len(df),
        'marcas': df['MARCA'].nunique() if 'MARCA' in df.columns else 0,
        'canales': df['CANAL'].nunique() if 'CANAL' in df.columns else 0,
        'columnas': list(df.columns)
    }


def _apply_excel_formatting(workbook):
    """Aplica formato a todas las hojas del workbook."""
    header_fill = PatternFill('solid', fgColor='7B68EE')  # Morado para TV
    header_font = Font(bold=True, color='FFFFFF')

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
